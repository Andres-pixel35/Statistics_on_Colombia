import os
import openpyxl
import pandas as pd

src = "../data/original/dane/poverty/poverty.xlsx"
out_dir = "../data/dane/poverty"


def cell(ws, r, c):
    return ws.cell(row=r, column=c).value


def to_num(s):
    """Comma-decimal strings ('9,6') and blank cells ('') show up alongside plain numbers."""
    return pd.to_numeric(s.map(lambda v: v.replace(",", ".") if isinstance(v, str) and v else v).replace("", pd.NA))


def read_domain_block(ws, header_row):
    """'Grandes dominios' row, year row below it, domain rows until a blank column A."""
    years = [v for v in next(ws.iter_rows(min_row=header_row + 1, max_row=header_row + 1, values_only=True))[1:] if v is not None]
    n = len(years)
    data = {}
    r = header_row + 2
    while cell(ws, r, 1) is not None:
        vals = [cell(ws, r, 2 + i) for i in range(n)]
        if any(v not in (None, "") for v in vals):
            data[cell(ws, r, 1)] = vals
        r += 1
    df = pd.DataFrame(data, index=pd.Index(years, name="Fecha"))
    return df.apply(to_num)


def read_category_block(ws, header_row, year):
    """'Características...' row, domain-columns header below it, category rows until Total."""
    domains = [v for v in next(ws.iter_rows(min_row=header_row + 1, max_row=header_row + 1, values_only=True))[2:] if v is not None]
    n = len(domains)
    rows = []
    grupo = None
    r = header_row + 2
    while True:
        label, sub = cell(ws, r, 1), cell(ws, r, 2)
        if label is not None:
            grupo = label
        vals = [cell(ws, r, 3 + i) for i in range(n)]
        rows.append([year, grupo, sub] + vals)
        if label == "Total":
            break
        r += 1
    df = pd.DataFrame(rows, columns=["Fecha", "Grupo", "Categoria"] + domains)
    df[domains] = df[domains].apply(to_num)
    return df


def read_sexo_block(ws, header_row):
    """'Grandes dominios' row holds 2 years, Hombre/Mujer row below it, domain rows below that."""
    year_row = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    gender_row = next(ws.iter_rows(min_row=header_row + 1, max_row=header_row + 1, values_only=True))
    year_by_col = {}
    last_year = None
    for c, v in enumerate(year_row[1:], start=2):
        if v is not None:
            last_year = v
        year_by_col[c] = last_year
    n = len(gender_row) - 1
    data = []
    r = header_row + 2
    while cell(ws, r, 1) is not None:
        domain = cell(ws, r, 1)
        vals = {2 + i: cell(ws, r, 2 + i) for i in range(n)}
        if any(v not in (None, "") for v in vals.values()):
            for c in range(2, 2 + n):
                data.append([year_by_col[c], gender_row[c - 1], domain, vals[c]])
        r += 1
    df = pd.DataFrame(data, columns=["Fecha", "Sexo", "Dominio", "Valor"])
    df["Valor"] = pd.to_numeric(df["Valor"].replace("", pd.NA))
    return df.pivot(index=["Fecha", "Sexo"], columns="Dominio", values="Valor").reset_index()


README = {
    "pobreza_monetaria": """# Pobreza Monetaria

`incidencia.csv` — incidence of monetary poverty, **% of population**, by domain (year rows).
`personas.csv` — people in monetary poverty, **thousands**, by domain (year rows).
""",
    "pobreza_extrema": """# Pobreza Extrema

`incidencia.csv` — incidence of extreme monetary poverty, **% of population**, by domain (year rows).
`personas.csv` — people in extreme monetary poverty, **thousands**, by domain (year rows).
""",
    "gini": """# Gini

`gini.csv` — Gini coefficient (0-1), by domain (year rows).
""",
    "ingreso_percapita": """# Ingreso per cápita

`ingreso_percapita.csv` — average per-capita household income, **current COP**, by domain (year rows).
""",
    "lineas_pobreza": """# Líneas de pobreza

`linea_pobreza.csv` — monetary poverty line, **current COP per person per month**, by domain (year rows).
`linea_pobreza_extrema.csv` — extreme poverty line, same units.
""",
    "brecha": """# Brecha de pobreza

`brecha_monetaria.csv` — poverty gap, **%**, by domain (year rows).
`brecha_extrema.csv` — extreme poverty gap, same units.
""",
    "severidad": """# Severidad de pobreza

`severidad_monetaria.csv` — poverty severity index, **%**, by domain (year rows).
`severidad_extrema.csv` — extreme poverty severity index, same units.
""",
    "perfil_jefe": """# Perfil del Jefe de Hogar

Poverty incidence (**%**) broken down by household-head characteristics (Sexo, Edad,
Nivel Educativo, Situación laboral, Posición Ocupacional, Cotización a pensión), plus a
`Total` row per year. Columns: `Fecha` (year), `Grupo`, `Categoria`, then one column per
domain (Nacional, Cabeceras, Centros poblados y rural disperso, 13 ciudades y A.M.,
Otras cabeceras).

- `incidencia_pobreza.csv` — monetary poverty.
- `incidencia_pobreza_extrema.csv` — extreme monetary poverty.
""",
    "perfil_hogar": """# Perfil del Hogar

Poverty incidence (**%**) broken down by household characteristics (Número de niños
menores de 12 años, Número de ocupados en el hogar, Tamaño del hogar), plus a `Total`
row per year. Same column layout as `perfil_jefe/`.

- `incidencia_pobreza.csv` — monetary poverty.
- `incidencia_pobreza_extrema.csv` — extreme monetary poverty.
""",
    "sexo": """# Pobreza según sexo

Poverty incidence (**%**) by sex of the person, by domain (incl. the 23 capital cities
and 7 metropolitan areas). Columns: `Fecha` (year), `Sexo`, then one column per domain.

- `incidencia_pobreza.csv` — monetary poverty.
- `incidencia_pobreza_extrema.csv` — extreme monetary poverty.
""",
}


def main():
    wb = openpyxl.load_workbook(src, data_only=True)

    outputs = {
        ("pobreza_monetaria", "incidencia.csv"): read_domain_block(wb["Pobreza Monetaria Act.Met."], 11),
        ("pobreza_monetaria", "personas.csv"): read_domain_block(wb["Pobreza Monetaria Act.Met."], 59),
        ("pobreza_extrema", "incidencia.csv"): read_domain_block(wb["Pobreza Extrema Act.Met."], 14),
        ("pobreza_extrema", "personas.csv"): read_domain_block(wb["Pobreza Extrema Act.Met."], 62),
        ("gini", "gini.csv"): read_domain_block(wb["Gini"], 11),
        ("ingreso_percapita", "ingreso_percapita.csv"): read_domain_block(wb["Ingpc Act.Met."], 11),
        ("lineas_pobreza", "linea_pobreza.csv"): read_domain_block(wb["LP_LI Act.Met."], 11),
        ("lineas_pobreza", "linea_pobreza_extrema.csv"): read_domain_block(wb["LP_LI Act.Met."], 50),
        ("brecha", "brecha_monetaria.csv"): read_domain_block(wb["Brecha Act.Met."], 11),
        ("brecha", "brecha_extrema.csv"): read_domain_block(wb["Brecha Act.Met."], 58),
        ("severidad", "severidad_monetaria.csv"): read_domain_block(wb["Severidad Act.Met."], 12),
        ("severidad", "severidad_extrema.csv"): read_domain_block(wb["Severidad Act.Met."], 59),
        ("perfil_jefe", "incidencia_pobreza.csv"): pd.concat([
            read_category_block(wb["IP_Perfil Jefe Act.Met."], 11, 2025),
            read_category_block(wb["IP_Perfil Jefe Act.Met."], 41, 2024),
        ], ignore_index=True),
        ("perfil_jefe", "incidencia_pobreza_extrema.csv"): pd.concat([
            read_category_block(wb["IPE_Perfil Jefe Act.Met."], 11, 2025),
            read_category_block(wb["IPE_Perfil Jefe Act.Met."], 41, 2024),
        ], ignore_index=True),
        ("perfil_hogar", "incidencia_pobreza.csv"): pd.concat([
            read_category_block(wb["IP_Perfil_Hogar Act.Met."], 11, 2025),
            read_category_block(wb["IP_Perfil_Hogar Act.Met."], 33, 2024),
        ], ignore_index=True),
        ("perfil_hogar", "incidencia_pobreza_extrema.csv"): pd.concat([
            read_category_block(wb["IPE_Perfil_Hogar Act.Met."], 11, 2025),
            read_category_block(wb["IPE_Perfil_Hogar Act.Met."], 33, 2024),
        ], ignore_index=True),
        ("sexo", "incidencia_pobreza.csv"): read_sexo_block(wb["IP_Sexo Act.Met."], 11),
        ("sexo", "incidencia_pobreza_extrema.csv"): read_sexo_block(wb["IPE_Sexo Act.Met."], 11),
    }

    for folder in {folder for folder, _ in outputs}:
        path = os.path.join(out_dir, folder)
        os.makedirs(path, exist_ok=True)
        with open(os.path.join(path, "README.md"), "w") as f:
            f.write(README[folder])

    for (folder, name), df in outputs.items():
        index = folder not in ("perfil_jefe", "perfil_hogar", "sexo")
        path = os.path.join(out_dir, folder, name)
        df.to_csv(path, index=index)
        print(f"{os.path.join(folder, name)}: {len(df)} rows")

    print(f"\nSaved {len(outputs)} CSVs to {out_dir}")


if __name__ == "__main__":
    main()
