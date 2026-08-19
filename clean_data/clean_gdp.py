import os
import openpyxl
import pandas as pd

PRODUCCION_PATH = "../data/original/dane/GDP/produccion.xlsx"
GASTO_PATH = "../data/original/dane/GDP/gasto.xlsx"
INGRESO_PATH = "../data/original/dane/GDP/ingreso.xlsx"
out_dir = "../data/dane/GDP"


def cell(ws, r, c):
    return ws.cell(row=r + 1, column=c + 1).value


def quarter_columns(ws, year_row, quarter_row, start_col):
    """(col, 'YYYY-Q') pairs, carrying the merged year label across its 4 quarter columns."""
    cols = []
    year = None
    c = start_col
    while True:
        y = cell(ws, year_row, c)
        if y is not None:
            year = str(int(y)) if isinstance(y, float) else str(y)
        q = cell(ws, quarter_row, c)
        if q is None:
            break
        cols.append((c, f"{year}-{q}"))
        c += 1
    return cols


def quarterly_table(ws, concept_col, year_row, quarter_row, row_range, start_col):
    """Concepto x quarter DataFrame, raw values (Miles de millones de pesos)."""
    cols = quarter_columns(ws, year_row, quarter_row, start_col)
    rows = []
    for r in row_range:
        concepto = cell(ws, r, concept_col)
        if concepto is None or str(concepto).strip() == "":
            continue
        values = [cell(ws, r, c) for c, _ in cols]
        rows.append([str(concepto).strip()] + values)
    columns = ["Concepto"] + [label for _, label in cols]
    return pd.DataFrame(rows, columns=columns)


def production(wb):
    ws = wb["Cuadro 3"]
    df = quarterly_table(ws, concept_col=3, year_row=11, quarter_row=12,
                          row_range=range(14, 103), start_col=4)
    return df.drop_duplicates(subset="Concepto", keep="first").reset_index(drop=True)


def spend_summarize(wb):
    ws = wb["Cuadro 1"]
    return quarterly_table(ws, concept_col=1, year_row=9, quarter_row=10,
                            row_range=range(12, 21), start_col=2)


def goal_homes_spend(wb):
    ws = wb["Cuadro 3"]
    return quarterly_table(ws, concept_col=1, year_row=9, quarter_row=10,
                            row_range=range(14, 27), start_col=2)


def durability_homes_spend(wb):
    ws = wb["Cuadro 3"]
    return quarterly_table(ws, concept_col=1, year_row=9, quarter_row=10,
                            row_range=range(30, 35), start_col=2)


def capital_formation(wb):
    ws = wb["Cuadro 5"]
    return quarterly_table(ws, concept_col=1, year_row=9, quarter_row=10,
                            row_range=range(12, 17), start_col=2)


def exports_and_imports(wb):
    ws = wb["Cuadro 7"]
    df = quarterly_table(ws, concept_col=1, year_row=9, quarter_row=10,
                          row_range=range(12, 18), start_col=2)
    # DANE reuses "Bienes"/"Servicios" under both Exportaciones and Importaciones
    df.loc[[1, 2], "Concepto"] = ["E.Bienes", "E.Servicios"]
    df.loc[[4, 5], "Concepto"] = ["I.Bienes", "I.Servicios"]
    return df


def income(wb):
    ws = wb["PIB_Ingreso"]
    cols = quarter_columns(ws, year_row=10, quarter_row=11, start_col=4)
    rows = []
    tag = None
    for r in range(13, 42):
        concepto = cell(ws, r, 2)
        if concepto is None or str(concepto).strip() == "":
            continue
        concepto = str(concepto).strip()
        if cell(ws, r, 1) is None:  # aggregate row, not an activity breakdown
            tag = "EBE" if concepto == "Excedente Bruto de Explotación" else \
                  "IM" if concepto == "Ingreso Mixto" else None
        else:
            concepto = f"{concepto} ({tag})"
        values = [cell(ws, r, c) for c, _ in cols]
        rows.append([concepto] + values)
    columns = ["Concepto"] + [label for _, label in cols]
    return pd.DataFrame(rows, columns=columns)


def main():
    prod_wb = openpyxl.load_workbook(PRODUCCION_PATH, data_only=True)
    gasto_wb = openpyxl.load_workbook(GASTO_PATH, data_only=True)
    ingreso_wb = openpyxl.load_workbook(INGRESO_PATH, data_only=True)

    outputs = {
        ("production", "summarize.csv"): production(prod_wb),
        ("spend", "summarize.csv"): spend_summarize(gasto_wb),
        ("spend", "goal_homes_spend.csv"): goal_homes_spend(gasto_wb),
        ("spend", "durability_homes_spend.csv"): durability_homes_spend(gasto_wb),
        ("spend", "capital_formation.csv"): capital_formation(gasto_wb),
        ("spend", "exports_and_imports.csv"): exports_and_imports(gasto_wb),
        ("income", "summarize.csv"): income(ingreso_wb),
    }

    for (folder, name), df in outputs.items():
        assert df["Concepto"].notna().all(), f"{folder}/{name}: null Concepto"
        assert not df.empty, f"{folder}/{name}: empty table"

    # production and spend are both real (constant prices); income is nominal
    # (precios corrientes), so its PIB isn't expected to match the other two
    prod_pib = outputs[("production", "summarize.csv")].set_index("Concepto").loc["Producto Interno Bruto"]
    spend_pib = outputs[("spend", "summarize.csv")].set_index("Concepto").loc["Producto Interno Bruto"]
    common = prod_pib.index.intersection(spend_pib.index)
    for q in common:
        assert prod_pib[q] == spend_pib[q], f"PIB mismatch at {q}: {prod_pib[q]} vs {spend_pib[q]}"

    for folder, _ in outputs:
        os.makedirs(os.path.join(out_dir, folder), exist_ok=True)

    for (folder, name), df in outputs.items():
        value_cols = [c for c in df.columns if c != "Concepto"]
        df[value_cols] = df[value_cols].round(0).astype("Int64")
        path = os.path.join(out_dir, folder, name)
        df.to_csv(path, index=False)
        print(f"{os.path.join(folder, name)}: {len(df)} rows")

    print(f"\nSaved {len(outputs)} CSVs to {out_dir}")


if __name__ == "__main__":
    main()
