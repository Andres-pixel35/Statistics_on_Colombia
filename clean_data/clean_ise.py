import os
import openpyxl
import pandas as pd

src = "../data/original/dane/ISE/anex-ISE-9actividades-abr2026.xlsx"
out_dir = "../data/dane/ISE"

MONTHS = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio",
          "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
START_YEAR = 2005

# (row offset from data_start, skip) for the 13 fixed data rows of each sub-table:
# 0=primarias header (dup of 1, skip), 1=primarias rama, 2=secundarias header (dup of 3, skip),
# 3=secundarias rama, 4=terciarias header (its own aggregate, keep), 5-11=terciarias ramas,
# 12=grand total (Indicador de Seguimiento a la Economía)
ROW_OFFSETS = [(0, True), (1, False), (2, True), (3, False), (4, False)] + \
    [(r, False) for r in range(5, 12)] + [(12, False)]


def cell(ws, r, c):
    return ws.cell(row=r + 1, column=c + 1).value


def read_table(ws, data_start):
    n_years = (ws.max_column - 1 + 11) // 12
    rows = []
    concepto = None
    for offset, skip in ROW_OFFSETS:
        r = data_start + offset
        label = cell(ws, r, 0)
        if offset in (0, 2, 4, 12):
            concepto = label
        if skip:
            continue
        for y in range(n_years):
            year = START_YEAR + y
            values = [cell(ws, r, 1 + y * 12 + m) for m in range(12)]
            if all(v is None for v in values):
                continue
            rows.append([year, concepto, label] + values)
    return pd.DataFrame(rows, columns=["Fecha", "Concepto", "Rama"] + MONTHS)


README = {
    "original": """# ISE - Datos originales (Cuadro 1)

Source: DANE, Indicador de Seguimiento a la Economia (ISE), 9-actividades breakdown.

- `indice.csv`: index points, base year 2015 = 100.
- `tasas_anuales.csv`: year-over-year growth rate (%).
- `tasas_ano_corrido.csv`: cumulative year-to-date growth rate (%).

Columns: `Fecha` (year), `Concepto` (Actividades primarias/secundarias/terciarias, or the
overall ISE total), `Rama` (specific activity, or same as Concepto for group/grand totals),
`Enero`..`Diciembre`.
""",
    "ajustado_estacional": """# ISE - Datos ajustados por estacionalidad y calendario (Cuadro 2)

Source: DANE, Indicador de Seguimiento a la Economia (ISE), 9-actividades breakdown.

- `indice.csv`: index points, base year 2015 = 100.
- `tasas_mensuales.csv`: month-over-month growth rate (%).
- `tasas_ano_corrido.csv`: cumulative year-to-date growth rate (%).

Columns: `Fecha` (year), `Concepto` (Actividades primarias/secundarias/terciarias, or the
overall ISE total), `Rama` (specific activity, or same as Concepto for group/grand totals),
`Enero`..`Diciembre`.
""",
}


def main():
    wb = openpyxl.load_workbook(src, data_only=True)

    original_dir = os.path.join(out_dir, "original")
    ajustado_dir = os.path.join(out_dir, "ajustado_estacional")

    ws1, ws2 = wb["Cuadro 1"], wb["Cuadro 2"]
    outputs = {
        (original_dir, "indice.csv"): read_table(ws1, 13),
        (original_dir, "tasas_anuales.csv"): read_table(ws1, 44),
        (original_dir, "tasas_ano_corrido.csv"): read_table(ws1, 75),
        (ajustado_dir, "indice.csv"): read_table(ws2, 13),
        (ajustado_dir, "tasas_mensuales.csv"): read_table(ws2, 44),
        (ajustado_dir, "tasas_ano_corrido.csv"): read_table(ws2, 75),
    }

    for folder in {original_dir, ajustado_dir}:
        os.makedirs(folder, exist_ok=True)

    with open(os.path.join(original_dir, "README.md"), "w") as f:
        f.write(README["original"])
    with open(os.path.join(ajustado_dir, "README.md"), "w") as f:
        f.write(README["ajustado_estacional"])

    for (folder, name), df in outputs.items():
        path = os.path.join(folder, name)
        df.to_csv(path, index=False)
        print(f"{os.path.join(os.path.basename(folder), name)}: {len(df)} rows")

    print(f"\nSaved {len(outputs)} CSVs to {out_dir}")


if __name__ == "__main__":
    main()
