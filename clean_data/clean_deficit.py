import os
import pandas as pd
import openpyxl

ANUAL_PATH = "../data/original/hacienda/deficit/balance_anual.xlsx"
MENSUAL_PATH = "../data/original/hacienda/deficit/balance_mensual.xlsx"
out_dir = "../data/hacienda/deficit"

HEADER_LABELS = {"Concepto", "CONCEPTO"}


def concepto_groups(concepts, levels):
    """Ancestor breadcrumb per row, from each row's outline indent level."""
    groups = []
    stack = []  # (level, label) of open ancestors
    for concepto, level in zip(concepts, levels):
        while stack and stack[-1][0] >= level:
            stack.pop()
        groups.append(" > ".join(label for _, label in stack))
        stack.append((level, concepto))
    return groups


def read_balance(xl, wb, sheet, year_cols=False):
    """Concepto rows x date columns: Anual/%PIB-anual/Mensual/Trimestral sheets."""
    df = pd.read_excel(xl, sheet_name=sheet, header=None)
    is_header_cell = lambda col: col.map(lambda v: isinstance(v, str) and v.strip() in HEADER_LABELS)
    # header cell sits in column 0 (Mensual/Trimestral) or column 1 (Anual, col 0 is a stray blank)
    hdr_col = next(c for c in (0, 1) if is_header_cell(df[c]).any())
    hdr = df.index[is_header_cell(df[hdr_col])][0]
    df.columns = df.iloc[hdr]
    df = df.loc[:, df.columns.notna()]  # drop the stray leading NaN-header column, if any
    df["_row"] = df.index  # raw 0-based row position, kept alive through the filters below
    df = df.iloc[hdr + 1:].reset_index(drop=True)
    df = df.rename(columns={df.columns[0]: "Concepto"})
    df["Concepto"] = df["Concepto"].str.strip()
    df = df[df["Concepto"].notna()].reset_index(drop=True)  # drop separator/artifact rows
    df = df[~df["Concepto"].str.startswith(("Fuente", "*"))].reset_index(drop=True)

    ws = wb[sheet]
    levels = [int(ws.cell(row=r + 1, column=hdr_col + 1).alignment.indent or 0) for r in df["_row"]]
    # The Mensual/Trimestral sheets indent "Funcionamiento*" one level too deep (same indent
    # as "Intereses deuda externa"), which would nest it and its children under Intereses.
    # The Anual sheet spells it "FUNCIONAMIENTO" (no asterisk) and is already correct, so this
    # comparison is a no-op there.
    levels = [2 if c.strip().rstrip("*") == "Funcionamiento" else lv
              for c, lv in zip(df["Concepto"], levels)]
    df["Grupo"] = concepto_groups(df["Concepto"], levels)
    df = df.drop(columns="_row")

    date_cols = [c for c in df.columns if c not in ("Concepto", "Grupo")]
    if year_cols:
        relabel = {c: (str(int(c)) if isinstance(c, (int, float)) else c.rstrip("*")) for c in date_cols}
    else:
        relabel = {c: c.strftime("%Y-%m-%d") for c in date_cols}
    df = df.rename(columns=relabel)
    date_cols = list(relabel.values())

    df[date_cols] = df[date_cols].replace("-", pd.NA).apply(pd.to_numeric)
    return df[["Concepto", "Grupo"] + date_cols]


README = {
    "anual": """# Anual

- `balance_cop.csv` — annual fiscal balance (Gobierno Nacional Central), **miles de millones
  (COP)** — same unit as `data/banco_republica/GDP/nominal_annual.csv`'s `PIB`.
- `balance_pib.csv` — same rows, **% of GDP** (DANE GDP denominator), already in
  percentage-point units (not a 0-1 fraction).

`Grupo` — the row's ancestor `Concepto` chain (e.g. `Renta`'s Grupo is
`1. INGRESOS TOTALES ... > INGRESOS CORRIENTES DE LA NACION > DIAN`), taken from the
source workbook's own outline indent levels. Empty for top-level rows
(`1. INGRESOS TOTALES`, `2. PAGOS TOTALES`, ...).

Both cover 1994-2025; **2025 figures are preliminary** per the source's `*Cifras
preliminares` footnote (the `*` marker was stripped from the column name).
""",
    "mensual": """# Mensual

- `balance_cop.csv` — monthly fiscal balance (Gobierno Nacional Central), **miles de millones
  (COP)** — same unit as `data/banco_republica/GDP/nominal_annual.csv`'s `PIB`.
- `balance_pib.csv` — same rows, **% of GDP**, already in percentage-point units
  (not a 0-1 fraction).

`Grupo` — the row's ancestor `Concepto` chain, taken from the source workbook's own
outline indent levels. Empty for top-level rows.
""",
    "trimestral": """# Trimestral

- `balance_cop.csv` — quarterly fiscal balance (Gobierno Nacional Central), **miles de millones
  (COP)** — same unit as `data/banco_republica/GDP/nominal_annual.csv`'s `PIB`.
- `balance_pib.csv` — same rows, **% of GDP**, already in percentage-point units
  (not a 0-1 fraction).

`Grupo` — the row's ancestor `Concepto` chain, taken from the source workbook's own
outline indent levels. Empty for top-level rows.
""",
}


def main():
    anual_xl, anual_wb = pd.ExcelFile(ANUAL_PATH), openpyxl.load_workbook(ANUAL_PATH)
    mensual_xl, mensual_wb = pd.ExcelFile(MENSUAL_PATH), openpyxl.load_workbook(MENSUAL_PATH)

    outputs = {
        ("anual", "balance_cop.csv"): read_balance(anual_xl, anual_wb, "Anual", year_cols=True),
        ("anual", "balance_pib.csv"): read_balance(anual_xl, anual_wb, "%PIB-DANE anual", year_cols=True),
        ("mensual", "balance_cop.csv"): read_balance(mensual_xl, mensual_wb, "Mensual ($MM)"),
        ("mensual", "balance_pib.csv"): read_balance(mensual_xl, mensual_wb, "Mensual (% del PIB)"),
        ("trimestral", "balance_cop.csv"): read_balance(mensual_xl, mensual_wb, "Trimestral ($MM)"),
        ("trimestral", "balance_pib.csv"): read_balance(mensual_xl, mensual_wb, "Trimestral (% del PIB)"),
    }

    for (folder, name), df in outputs.items():
        assert df["Concepto"].notna().all(), f"{folder}/{name}: null Concepto"
        value_cols = [c for c in df.columns if c not in ("Concepto", "Grupo")]
        assert df[value_cols].apply(lambda c: pd.api.types.is_numeric_dtype(c)).all(), (
            f"{folder}/{name}: non-numeric value column"
        )
        top_level = df["Concepto"].str.match(r"^\d\.")
        assert (df.loc[top_level, "Grupo"] == "").all(), f"{folder}/{name}: top-level row has a non-empty Grupo"
    for grain in ("anual", "mensual", "trimestral"):
        cop = outputs[(grain, "balance_cop.csv")]
        pib = outputs[(grain, "balance_pib.csv")]
        # row-count parity only: source sheets occasionally reword a Concepto slightly
        # between the COP and %PIB versions of the same table (e.g. the Anual sheets'
        # row 41), so exact label equality isn't a reliable invariant
        assert len(cop) == len(pib), f"{grain}: balance_cop.csv/balance_pib.csv row count mismatch"

    for folder in {folder for folder, _ in outputs}:
        os.makedirs(os.path.join(out_dir, folder), exist_ok=True)
        with open(os.path.join(out_dir, folder, "README.md"), "w") as f:
            f.write(README[folder])

    for (folder, name), df in outputs.items():
        path = os.path.join(out_dir, folder, name)
        df.to_csv(path, index=False)
        print(f"{os.path.join(folder, name)}: {len(df)} rows")

    print(f"\nSaved {len(outputs)} CSVs to {out_dir}")


if __name__ == "__main__":
    main()
