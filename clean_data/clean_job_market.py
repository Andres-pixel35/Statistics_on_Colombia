import os
import re
import openpyxl
import pandas as pd
from generalities.function import norm

src_dir = "./data/original/dane/job_market"
out_dir = "./data/dane/job_market"

# file -> dimension column name, output sub-folder, {sheet_index: csv_name}
FILES = [
    {"file": "mercado_laboral.xlsx", "dim": "Perspectiva", "folder": "Mercado Laboral",
     "sheets": {3: "total", 6: "posicion_ocupacional", 7: "ramas_actividad",
                8: "fuera_fuerza_trabajo"}},
    {"file": "departamentos.xls", "dim": "Departamentos", "folder": "Departamentos",
     "sheets": {2: "total", 3: "hombres", 4: "mujeres", 7: "ramas_actividad"}},
    {"file": "informalidad.xlsx", "dim": "Perspectiva", "folder": "informalidad",
     "sheets": {6: "total", 8: "sexo", 9: "educacion", 10: "ramas_actividad",
                11: "posicion_ocupacional", 12: "lugar_trabajo", 13: "tamano_empresa",
                14: "seguridad_social", 15: "seguridad_social_sexo"}},
    {"file": "regiones.xls", "dim": "Perspectiva", "folder": "regiones",
     "sheets": {3: "total", 4: "sexo"}},
    {"file": "infantil.xlsx", "dim": "Perspectiva", "folder": "infantil",
     "sheets": {2: "total", 3: "sexo", 4: "edad", 6: "asistencia_escolar",
                7: "razon", 8: "horas", 9: "rama_actividad", 10: "posicion",
                11: "ingreso", 13: "actividades_sexo"}},
]

GENDER = {"hombres": "Hombres", "mujeres": "Mujeres",
          "hombre": "Hombres", "mujer": "Mujeres"}
YEAR_RE = re.compile(r"(19|20)\d{2}")
MONTH_SEQ = ["ene", "feb", "mar", "abr", "may", "jun",
             "jul", "ago", "sep", "oct", "nov", "dic"]
MONTHS = set(MONTH_SEQ)
MONTH_ORD = {m: i + 1 for i, m in enumerate(MONTH_SEQ)}
GENDER_SUFFIX = re.compile(r"^(.*\S)\s*-\s*(hombres|mujeres)\s*$", re.I)

# Sheets where the whole sheet is one gender (block titles are departments).
FILE_SEXO = {("departamentos.xls", 3): "Hombres",
             ("departamentos.xls", 4): "Mujeres"}


def read_sheet(path, sheet_idx):
    engine = "openpyxl" if path.lower().endswith(".xlsx") else "xlrd"
    return pd.read_excel(path, sheet_name=sheet_idx, header=None, engine=engine)


def read_bold(path, sheet_idx):
    """0-based row indices whose col-0 cell is bold (group headers).

    Only .xlsx carries readable formatting (xlrd 2.x drops it); the .xls files
    are flat, so they return an empty set and gain no grouping.
    """
    if not path.lower().endswith(".xlsx"):
        return frozenset()
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[sheet_idx]
    bold = {r for r in range(ws.max_row)
            if ws.cell(row=r + 1, column=1).font.bold
            and ws.cell(row=r + 1, column=1).value not in (None, "")}
    return bold


def is_year(v):
    try:
        f = float(v)
    except (TypeError, ValueError):
        return False
    return 1990 <= f <= 2100 and f == int(f)


def is_num(v):
    return pd.notna(pd.to_numeric(v, errors="coerce"))


def normalize_period(label):
    """Collapse DANE period variants to one label.

    Strips embedded years and footnote marks, standardizes hyphen spacing, and
    capitalizes month tokens so 'Nov 07 - ene 08' -> 'Nov - Ene', 'IV - 2022' ->
    'IV', 'Ene - mar'/'Ago- Oct'/'Feb-Abr'/'Ene - Mar*' all fold onto one label.
    Roman numerals ('IV', 'I', 'II') and 'Anual' are left untouched.
    """
    s = re.sub(r"\d+", "", str(label)).replace("*", "").replace("^", "")
    s = re.sub(r"\s*-\s*", " - ", s)
    s = re.sub(r"\s+", " ", s).strip().strip(" -").strip()
    return " ".join(t.capitalize() if t.lower() in MONTHS else t for t in s.split(" "))


def month_index(periodo):
    """1-based index of a normalized period's first month ('Ene - Mar' -> 1), else None."""
    tok = str(periodo).split(" ")[0].lower()[:3]
    return MONTH_ORD.get(tok)


def period_start_year(label):
    """First year token in a cross-year period label ('Nov 25 - ene 26' -> 2025).

    DANE embeds the start month's year only on cross-year quarters; plain labels
    ('Ene - mar') carry no digits and return None so the caller keeps the header year.
    Two-digit tokens are 2000-relative.
    """
    m = re.search(r"\b(\d{2,4})\b", str(label))
    if not m:
        return None
    y = int(m.group(1))
    return y if y >= 1990 else 2000 + y


def title_above(df, anchor):
    """Nearest non-null col0 text above the anchor row = the block title."""
    for r in range(anchor - 1, -1, -1):
        cell = df.iloc[r, 0]
        if pd.notna(cell) and str(cell).strip():
            return str(cell).strip()
    return ""


def split_gender(title):
    """'Total Nacional - Hombres' -> ('Total Nacional', 'Hombres').

    Standalone 'HOMBRES' (no dash) returns (title, None) so it stays a block title.
    """
    m = GENDER_SUFFIX.match(title)
    if m:
        return m.group(1).strip(), GENDER[m.group(2).lower()]
    return title, None


def classify(df, H):
    """Return (data_start, {col: (year, periodo)}) for the block at anchor row H.

    Four families: C embedded 'IV - YYYY' headers; A year row + period row;
    annual with years on the Concepto row; annual with years on the next row.
    """
    ncols = df.shape[1]
    row_h = df.iloc[H]
    row_h1 = df.iloc[H + 1] if H + 1 < len(df) else None
    cols = range(1, ncols)

    # Family C: header strings embed a period token AND a year ('IV - 2022').
    # A bare text-typed year like '2008' (normalize_period -> '') is NOT family C —
    # it is a family-A year cell that happens to be stored as text.
    cc = [c for c in cols if isinstance(row_h[c], str)
          and YEAR_RE.search(row_h[c]) and normalize_period(row_h[c])]
    if cc:
        return H + 1, {c: (int(YEAR_RE.search(row_h[c]).group(0)),
                           normalize_period(row_h[c])) for c in cc}

    years_h = [c for c in cols if is_year(row_h[c])]
    if years_h:
        text_h1 = (row_h1 is not None and
                   any(isinstance(row_h1[c], str) and row_h1[c].strip()
                       and not is_year(row_h1[c]) for c in cols))
        if text_h1:  # Family A: years on H (may have gaps), period labels on H+1
            colmap, cur, prev_m = {}, None, None
            for c in cols:
                explicit = is_year(row_h[c])
                if explicit:
                    cur = int(row_h[c])
                lab = row_h1[c]
                if not (isinstance(lab, str) and lab.strip()):
                    continue
                periodo = normalize_period(lab)
                sy = period_start_year(lab)
                m = month_index(periodo)
                if sy is not None:  # cross-year label carries its own start year
                    cur = sy
                elif (not explicit and cur is not None and prev_m is not None
                      and m is not None and m < prev_m):
                    cur += 1  # quarters wrapped past December -> new year (fills header gaps)
                if cur is not None:
                    colmap[c] = (cur, periodo)
                if m is not None:
                    prev_m = m
            return H + 2, colmap
        # Annual: years on the Concepto row, no period row
        return H + 1, {c: (int(row_h[c]), "Anual") for c in years_h}

    # Annual: years on the row below the Concepto row
    if row_h1 is not None:
        years_h1 = [c for c in cols if is_year(row_h1[c])]
        if years_h1:
            return H + 2, {c: (int(row_h1[c]), "Anual") for c in years_h1}
    return H + 1, {}


def section_map(df, anchors, titles, bold_rows):
    """Map each Concepto anchor to its section label, or None if the sheet is flat.

    Some sheets stack two tables under the same block titles (e.g. absolute counts
    then 'Distribución porcentual ...'), distinguished only by a bold, value-less
    divider sitting *between* blocks. A divider is a bold col-0 row that is neither a
    'Concepto' row nor a block title (those repeat per block). Sections activate only
    when such a divider appears after the first anchor; otherwise every anchor maps to
    None and grouping falls back to the per-block logic. The initial (pre-divider)
    section is named from the longest divider candidate in the top matter (the metric
    title), so the first stacked table gets a meaningful, distinct label too.
    """
    if not anchors:
        return {}
    ncols = df.shape[1]
    block_titles = {t for t in titles.values() if t}

    def has_value(r):
        return any(is_num(df.iloc[r, c]) for c in range(1, ncols))

    def candidate(r):  # bold, value-less, not a 'Concepto' row, not a block title
        c0 = df.iloc[r, 0]
        if pd.isna(c0) or not str(c0).strip() or r not in bold_rows:
            return None
        label = str(c0).strip()
        if norm(label) == "concepto" or label in block_titles or has_value(r):
            return None
        return label

    def next_title_row(r):  # title row of the first block starting after r, else None
        a = next((a for a in anchors if a > r), None)
        if a is None:
            return None
        return next((rr for rr in range(a - 1, -1, -1)
                     if pd.notna(df.iloc[rr, 0]) and str(df.iloc[rr, 0]).strip()), a)

    def is_divider(r):
        # A section divider heads a NEW block stack: it has no valued rows between it
        # and the next block's title (a subgroup header is followed by its members).
        if not candidate(r):
            return False
        t = next_title_row(r)
        return t is not None and r < t and not any(has_value(rr)
                                                   for rr in range(r + 1, t))

    first = anchors[0]
    inter = [r for r in range(first + 1, len(df)) if is_divider(r)]
    if not inter:
        return {H: None for H in anchors}

    top = [lbl for r in range(first) if (lbl := candidate(r))]
    current = max(top, key=len, default=None)
    result, ai = {}, 0
    for r in range(len(df)):
        if r in inter:
            current = candidate(r)
        if ai < len(anchors) and r == anchors[ai]:
            result[anchors[ai]] = current
            ai += 1
    return result


def parse_sheet(df, dim, default_sexo="Total", bold_rows=frozenset()):
    anchors = [r for r in range(len(df)) if norm(df.iloc[r, 0]) == "concepto"]
    titles = {a: title_above(df, a) for a in anchors}
    sections = section_map(df, anchors, titles, bold_rows)
    records = []

    for i, H in enumerate(anchors):
        # Gender may live in the title ('Total Nacional - Hombres'); split it out.
        perspectiva, title_sexo = split_gender(titles[H])
        data_start, colmap = classify(df, H)
        if not colmap:
            continue
        end = titles_index(anchors, titles, i, df)
        section = sections.get(H)  # stacked-table label, or None for flat sheets

        def emit(row, concept, sexo, grupo):
            for c, (year, periodo) in colmap.items():
                num = pd.to_numeric(df.iloc[row, c], errors="coerce")
                if pd.notna(num):
                    records.append({"Fecha": year, dim: perspectiva, "Sexo": sexo,
                                    "Grupo": grupo, "Concepto": concept,
                                    "Periodo": periodo, "Valor": float(num)})

        # Blank-row subgroups only disambiguate when a block holds >=2 of them
        # (e.g. asistencia's total vs working population); a lone headline is noise.
        subgroups, pend = 0, True
        for r in range(data_start, end):
            c0 = df.iloc[r, 0]
            if pd.isna(c0) or not str(c0).strip():
                pend = True
                continue
            label = str(c0).strip()
            if norm(label) == "concepto" or norm(label) in GENDER:
                continue
            if r in bold_rows:
                pend = False
                continue
            if pend and any(is_num(df.iloc[r, c]) for c in colmap):
                subgroups += 1
                pend = False
        use_subgroups = subgroups >= 2

        sexo = title_sexo or default_sexo
        group = None  # bold formality header (Población ocupada / Formal / Informal)
        headline = None  # first normal concept of the block (e.g. 'Población ocupada')
        pending = True  # expect a (sub)group headline; reset after each blank row
        for r in range(data_start, end):
            c0 = df.iloc[r, 0]
            if pd.isna(c0) or not str(c0).strip():
                pending = True  # blank row separates subgroups within a block
                continue
            label = str(c0).strip()
            if norm(label) == "concepto":
                continue
            grupo = section if section is not None else group
            if norm(label) in GENDER:  # nested gender sub-header (valued or not)
                sexo = GENDER[norm(label)]
                if headline is not None and any(is_num(df.iloc[r, c]) for c in colmap):
                    emit(r, headline, sexo, grupo)  # valued header = gender's headline total
                continue
            valued = any(is_num(df.iloc[r, c]) for c in colmap)
            if r in bold_rows:  # bold concept = group header for the rows below it
                group, pending = label, False
            elif use_subgroups and pending and valued:  # non-bold subgroup headline
                group, pending = label, False
            grupo = section if section is not None else group
            if not valued:
                continue
            if headline is None:
                headline = label
            emit(r, label, sexo, grupo)

    return pd.DataFrame(records, columns=["Fecha", dim, "Sexo", "Grupo",
                                          "Concepto", "Periodo", "Valor"])


def titles_index(anchors, titles, i, df):
    """Block data ends at the title row of the next block (or end of sheet)."""
    if i + 1 < len(anchors):
        next_anchor = anchors[i + 1]
        for r in range(next_anchor - 1, -1, -1):
            cell = df.iloc[r, 0]
            if pd.notna(cell) and str(cell).strip():
                return r
        return next_anchor
    return len(df)


def drop_dead_columns(out):
    """Drop Grupo unless it disambiguates, then drop any single-value column.

    Grupo is kept only when some Concepto repeats under >=2 distinct groups
    (the real ambiguity it solves); flat sheets lose it. Sexo/Periodo (and a
    surviving constant Grupo) are dropped when they carry a single value.
    """
    distinct_groups = out.groupby("Concepto")["Grupo"].nunique(dropna=True)
    if (distinct_groups >= 2).any():
        out["Grupo"] = out["Grupo"].fillna("")
    else:
        out = out.drop(columns="Grupo")
    dead = [c for c in ("Sexo", "Grupo", "Periodo")
            if c in out.columns and out[c].nunique(dropna=False) <= 1]
    return out.drop(columns=dead)


def main():
    total_csv = 0
    for cfg in FILES:
        path = os.path.join(src_dir, cfg["file"])
        folder = os.path.join(out_dir, cfg["folder"])
        os.makedirs(folder, exist_ok=True)
        for idx, name in cfg["sheets"].items():
            df = read_sheet(path, idx)
            default_sexo = FILE_SEXO.get((cfg["file"], idx), "Total")
            out = parse_sheet(df, cfg["dim"], default_sexo, read_bold(path, idx))
            out = drop_dead_columns(out)
            out.to_csv(os.path.join(folder, name + ".csv"), index=False)
            total_csv += 1
            print(f"{cfg['folder']}/{name}.csv: {len(out)} rows")
    print(f"\nSaved {total_csv} CSVs to {out_dir}")


if __name__ == "__main__":
    main()
