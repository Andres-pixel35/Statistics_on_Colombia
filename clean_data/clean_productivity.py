import os
import re
import openpyxl
import pandas as pd

src_dir = "../data/original/dane/productivity"
out_dir = "../data/dane/productivity"

YEAR_RE = re.compile(r"(\d{4})")


def cell(ws, r, c):
    return ws.cell(row=r + 1, column=c + 1).value


def year_int(v):
    return int(YEAR_RE.match(str(v)).group(1))


def read_rows(ws, row_start, year_col):
    r = row_start
    while cell(ws, r, year_col) is not None:
        yield r
        r += 1


def valor_agregado_total_factores(wb):
    """National annual PTF series, Enfoque Valor Agregado (Productividad workbook Cuadro 1)."""
    ws = wb["Cuadro 1"]
    rows = [[year_int(cell(ws, r, 1))] + [cell(ws, r, c) for c in range(2, 7)]
            for r in read_rows(ws, 15, 1)]
    cols = ["año", "valor agregado bruto (%)", "servicios laborales (pp)",
            "servicios de capital (pp)", "contribución de los factores (pp)",
            "productividad total de los factores (pp)"]
    return pd.DataFrame(rows, columns=cols)


def nacional_cuadro3(wb):
    """National annual PTF, Enfoque Producción (Cuadro 3)."""
    ws = wb["Cuadro 3"]
    rows = [[year_int(cell(ws, r, 1))] + [cell(ws, r, c) for c in range(2, 10)]
            for r in read_rows(ws, 15, 1)]
    cols = ["año", "producción (%)", "servicios laborales mujeres (pp)",
            "servicios laborales hombres (pp)", "servicios laborales total (pp)",
            "servicios de capital (pp)", "consumos intermedios (pp)",
            "contribución de los factores (pp)", "productividad total de los factores (pp)"]
    return pd.DataFrame(rows, columns=cols)


def actividad_economica_families(wb):
    """Per-actividad-económica PTF breakdown across all years (Cuadros 4-23, one per year)."""
    rows = []
    for idx in range(4, 24):
        ws = wb[f"Cuadro {idx}"]
        year = year_int(cell(ws, 10, 0))
        for r in read_rows(ws, 15, 1):
            actividad = cell(ws, r, 1)
            rows.append([year, actividad] + [cell(ws, r, c) for c in range(2, 15)])

    cols = ["año", "Actividad Económica", "Tasas de crecimiento (%)",
            "Composición del trabajo (pp)", "Horas trabajadas (pp)", "Laboral Total (pp)",
            "Capital TIC (pp)", "Capital No TIC (pp)", "Capital Total (pp)",
            "Energía (pp)", "Materiales (pp)", "Servicios (pp)", "Consumo intermedio Total (pp)",
            "Contribución de los factores (pp)", "Productividad Total de los Factores (pp)"]
    return pd.DataFrame(rows, columns=cols)


def laboral_cuadro(ws, value_cols, extra_cols):
    """National annual labor-productivity series (ProductividadLaboral workbook, one sheet)."""
    rows = [[year_int(cell(ws, r, 1))] + [cell(ws, r, c) for c in value_cols]
            for r in read_rows(ws, 14, 1)]
    return pd.DataFrame(rows, columns=["año"] + extra_cols)


def main():
    wb = openpyxl.load_workbook(os.path.join(src_dir, "anex-PTF-Productividad-2025.xlsx"),
                                 data_only=True)
    wb_laboral = openpyxl.load_workbook(
        os.path.join(src_dir, "anex-PTF-ProductividadLaboral-2025.xlsx"), data_only=True)

    outputs = {}

    valor_agregado_dir = os.path.join(out_dir, "valor_agregado")
    outputs[(valor_agregado_dir, "productividad_total_factores.csv")] = \
        valor_agregado_total_factores(wb)

    nacional_dir = os.path.join(out_dir, "produccion", "nacional")
    outputs[(nacional_dir, "nacional.csv")] = nacional_cuadro3(wb)

    actividad_dir = os.path.join(out_dir, "produccion", "actividad_economica")
    outputs[(actividad_dir, "actividad_economica.csv")] = actividad_economica_families(wb)

    laboral_dir = os.path.join(out_dir, "laboral")
    outputs[(laboral_dir, "por_hora_trabajada.csv")] = laboral_cuadro(
        wb_laboral["Cuadro 1"], range(2, 6),
        ["productividad laboral por hora trabajada (%)",
         "productividad total de los factores (pp)", "composición laboral (pp)",
         "contribución del capital por hora trabajada (pp)"])
    outputs[(laboral_dir, "por_persona_empleada.csv")] = laboral_cuadro(
        wb_laboral["Cuadro 2"], range(2, 7),
        ["productividad laboral por persona empleada (%)",
         "productividad total de los factores (pp)",
         "contribución de las horas trabajadas por empleado (pp)",
         "composición laboral (pp)",
         "contribución del capital por persona empleada (pp)"])

    for folder in {valor_agregado_dir, nacional_dir, actividad_dir, laboral_dir}:
        os.makedirs(folder, exist_ok=True)

    for (folder, name), df in outputs.items():
        path = os.path.join(folder, name)
        df.to_csv(path, index=False)
        print(f"{os.path.join(os.path.basename(folder), name)}: {len(df)} rows")

    print(f"\nSaved {len(outputs)} CSVs to {out_dir}")


if __name__ == "__main__":
    main()
